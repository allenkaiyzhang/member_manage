import pandas
from openpyxl import load_workbook
#edit the information of members
#TODO:add the lowest nember 

def update_member(id,name,sex,age):

        wb = load_workbook(r'.\\member_info_list.xlsx')
        sheet = wb.activesheet.cell(row=i, column="A").value = id
        sheet = wb.activesheet.cell(row=i, column="B").value = name
        sheet = wb.activesheet.cell(row=i, column="C").value = sex
        sheet = wb.activesheet.cell(row=i, column="D").value = age
        wb.save(r'.\\member_info_list.xlsx')
        pass

def add_number(id,name,sex,age):

        wb = load_workbook(r'.\\member_info_list.xlsx')
        sheet = wb.activesheet.cell(row=i, column="A").value = id
        sheet = wb.activesheet.cell(row=i, column="B").value = name
        sheet = wb.activesheet.cell(row=i, column="C").value = sex
        sheet = wb.activesheet.cell(row=i, column="D").value = age
        wb.save(r'.\\member_info_list.xlsx')
        pass


def del_member(name):
    
        wb = load_workbook(r'.\\member_info_list.xlsx')
        sheet = wb.activesheet.cell(row=i, column="A").value=""
        sheet = wb.activesheet.cell(row=i, column="B").value=""
        sheet = wb.activesheet.cell(row=i, column="C").value =""
        sheet = wb.activesheet.cell(row=i, column="D").value =""
        wb.save(r'.\\member_info_list.xlsx')
        pass

def test():
    print("Hello,World!")
    
